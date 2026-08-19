"""Contact import: parse, map, normalise, dedupe, and store (FRD §6).

A list belongs to the agent it was mapped for (FR-CONTACT-1); parsing,
header detection and normalisation all happen here, server side
(FR-CONTACT-2). The uploaded file is the source of truth — every
mapping change recomputes the stored rows from the original bytes, so
nothing accumulates from stale mappings.

The mapping dict maps a spreadsheet column NAME to a field id, the
string "phone", or None (FR-DATA-4: ids, never field names).
"""

import csv
import io
import re
import uuid
from dataclasses import dataclass
from difflib import SequenceMatcher
from typing import Any

import phonenumbers
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from becca.domain.dedupe import dedupe_key
from becca.domain.model import Field

ColumnTarget = int | str | None  # field id | "phone" | unused
Mapping = dict[str, ColumnTarget]


class UnreadableFile(Exception):
    """The upload could not be parsed as CSV or XLSX at all."""


@dataclass(frozen=True)
class ParsedFile:
    headers: tuple[str, ...]
    rows: tuple[tuple[str, ...], ...]


# ---------------------------------------------------------------- parsing


def _cell_str(value: Any) -> str:
    """A cell as the text a user would consider it to hold.

    Excel types phone columns as numbers and dates as datetimes; str()
    would yield "8030001188.0" and "2026-08-07 00:00:00" — artifacts
    the sheet never displayed and the agent should never speak.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if hasattr(value, "isoformat"):
        iso: str = value.isoformat()
        return iso.removesuffix("T00:00:00").replace("T", " ")
    return str(value).strip()


def _finish(raw_rows: list[list[str]]) -> ParsedFile:
    rows = [r for r in raw_rows if any(c.strip() for c in r)]
    if not rows:
        raise UnreadableFile("no rows")
    headers = [h.strip() for h in rows[0]]
    width = len(headers)
    # Unnamed or repeated headers still need distinct mapping keys.
    seen: dict[str, int] = {}
    for i, h in enumerate(headers):
        name = h or f"Column {i + 1}"
        seen[name] = seen.get(name, 0) + 1
        headers[i] = name if seen[name] == 1 else f"{name} ({seen[name]})"
    data = [tuple((r[i].strip() if i < len(r) else "") for i in range(width)) for r in rows[1:]]
    if not data:
        raise UnreadableFile("headers but no contact rows")
    return ParsedFile(headers=tuple(headers), rows=tuple(data))


def _parse_csv(data: bytes) -> ParsedFile:
    try:
        text_data = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text_data = data.decode("cp1252", errors="replace")
    try:
        dialect: csv.Dialect | type[csv.Dialect] = csv.Sniffer().sniff(
            text_data[:4096], delimiters=",;\t"
        )
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text_data), dialect)
    return _finish([[c for c in row] for row in reader])


def _parse_xlsx(data: bytes) -> ParsedFile:
    from openpyxl.reader.excel import load_workbook

    try:
        book = load_workbook(  # type: ignore[no-untyped-call]  # openpyxl ships no types
            io.BytesIO(data), read_only=True, data_only=True
        )
    except Exception as exc:
        raise UnreadableFile(str(exc)) from exc
    try:
        sheet = book.worksheets[0]
        return _finish([[_cell_str(c) for c in row] for row in sheet.iter_rows(values_only=True)])
    finally:
        book.close()


def parse_upload(filename: str, data: bytes) -> ParsedFile:
    """CSV or XLSX (FR-CONTACT-2), decided by extension with a content
    fallback: xlsx files are zip archives, so a PK signature wins over a
    misleading name."""
    if filename.lower().endswith(".xlsx") or data[:2] == b"PK":
        return _parse_xlsx(data)
    return _parse_csv(data)


# ------------------------------------------------------------- suggestion

_PHONE_HEADERS = {
    "phone",
    "phonenumber",
    "phoneno",
    "telephone",
    "tel",
    "mobile",
    "mobileno",
    "mobilenumber",
    "msisdn",
    "gsm",
    "gsmno",
    "number",
    "contactnumber",
    "whatsapp",
}


def _norm(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", name.lower())


def _tokens(name: str) -> frozenset[str]:
    return frozenset(t for t in re.split(r"[^a-z0-9]+", name.lower()) if t)


def _similarity(header: str, field_key: str) -> float:
    """Deliberately generous: a suggestion is only a suggestion — the
    user confirms every mapping (FR-CONTACT-4). "Full Name" should
    reach first_name; word overlap catches what edit distance misses."""
    a, b = _norm(header), _norm(field_key)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.9
    key_tokens = _tokens(field_key)
    token_overlap = len(_tokens(header) & key_tokens) / len(key_tokens) if key_tokens else 0.0
    return max(SequenceMatcher(None, a, b).ratio(), token_overlap)


def suggest_mapping(headers: tuple[str, ...], contract: tuple[Field, ...]) -> Mapping:
    """Auto-suggest by name similarity, confirmed by the user
    (FR-CONTACT-4). Greedy best-match; a column serves one target."""
    mapping: Mapping = {h: None for h in headers}
    taken: set[str] = set()
    phone = next((h for h in headers if _norm(h) in _PHONE_HEADERS), None)
    if phone is not None:
        mapping[phone] = "phone"
        taken.add(phone)
    candidates = sorted(
        ((_similarity(h, f.key), h, f.id) for f in contract for h in headers if h not in taken),
        key=lambda t: t[0],
        reverse=True,
    )
    mapped_fields: set[int] = set()
    for score, header, field_id in candidates:
        if score < 0.5:
            break
        if header in taken or field_id in mapped_fields:
            continue
        mapping[header] = field_id
        taken.add(header)
        mapped_fields.add(field_id)
    return mapping


def unmapped_required(contract: tuple[Field, ...], mapping: Mapping) -> tuple[Field, ...]:
    """Required input fields no column maps to — each one blocks
    (FR-CONTACT-5), remedied by mapping a column or making the field
    optional."""
    mapped_ids = {v for v in mapping.values() if isinstance(v, int)}
    return tuple(f for f in contract if f.required and f.id not in mapped_ids)


def phone_column(mapping: Mapping) -> str | None:
    return next((h for h, v in mapping.items() if v == "phone"), None)


# ---------------------------------------------------------- normalisation


def normalise_phone(raw: str) -> str | None:
    """E.164 or None (FR-CONTACT-3). Region NG: v1 is single-country."""
    if not raw.strip():
        return None
    try:
        parsed = phonenumbers.parse(raw, "NG")
    except phonenumbers.NumberParseException:
        return None
    if not phonenumbers.is_valid_number(parsed):
        return None
    result: str = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
    return result


# ------------------------------------------------------------ computation


@dataclass(frozen=True)
class ComputedRow:
    row_index: int  # 1-based position among the file's data rows
    phone_raw: str
    phone_e164: str | None
    variables: dict[str, str]  # field id (as string) -> value
    dedupe_key: str
    diallable: bool
    exclusion_reason: str | None


@dataclass(frozen=True)
class ComputedImport:
    rows: tuple[ComputedRow, ...]  # duplicates already dropped
    duplicate_count: int

    @property
    def diallable_count(self) -> int:
        return sum(1 for r in self.rows if r.diallable)

    def excluded(self, reason: str) -> int:
        return sum(1 for r in self.rows if r.exclusion_reason == reason)


def compute_import(
    parsed: ParsedFile, contract: tuple[Field, ...], mapping: Mapping
) -> ComputedImport:
    """The whole row pipeline: normalise, key, dedupe, completeness.

    Normalisation runs before the key so 0803… and +234803… collide
    (FR-CONTACT-8); first occurrence wins. A row failing E.164 keys on
    its raw value — identical unreadable rows still collapse. Required
    completeness is judged over MAPPED required fields only: an entirely
    unmapped required field is the list-level block of FR-CONTACT-5,
    not a per-row defect.
    """
    col_index = {h: i for i, h in enumerate(parsed.headers)}
    phone_col = phone_column(mapping)
    field_cols = [(col_index[h], v) for h, v in mapping.items() if isinstance(v, int)]
    mapped_required = {
        f.id for f in contract if f.required and f.id in {fid for _, fid in field_cols}
    }

    seen: set[str] = set()
    rows: list[ComputedRow] = []
    duplicates = 0
    for i, cells in enumerate(parsed.rows, start=1):
        phone_raw = cells[col_index[phone_col]] if phone_col is not None else ""
        e164 = normalise_phone(phone_raw)
        variables = {str(fid): cells[ci] for ci, fid in field_cols if cells[ci].strip()}
        key = dedupe_key(e164 or phone_raw, variables)
        if key in seen:
            duplicates += 1
            continue
        seen.add(key)
        if e164 is None:
            reason = "unparseable_number"
        elif any(str(fid) not in variables for fid in mapped_required):
            reason = "missing_required_value"
        else:
            reason = None
        rows.append(
            ComputedRow(
                row_index=i,
                phone_raw=phone_raw,
                phone_e164=e164,
                variables=variables,
                dedupe_key=key,
                diallable=reason is None,
                exclusion_reason=reason,
            )
        )
    return ComputedImport(rows=tuple(rows), duplicate_count=duplicates)


def resolved_variables(contract: tuple[Field, ...], variables: dict[str, str]) -> dict[int, str]:
    """A row's spoken values: stored values, optionals falling back to
    their natural-reading default (FR-CONTACT-9). Keyed by field id;
    dispatch uses this at dial time."""
    out: dict[int, str] = {}
    for f in contract:
        value = variables.get(str(f.id), "").strip()
        if not value and not f.required:
            value = f.default
        if value:
            out[f.id] = value
    return out


# ---------------------------------------------------------------- storage


def _mapping_to_json(mapping: Mapping) -> str:
    import json

    return json.dumps(mapping)


def _mapping_from_json(data: Any) -> Mapping:
    import json

    raw: dict[str, Any] = data if isinstance(data, dict) else json.loads(data)
    return {
        h: (int(v) if isinstance(v, int | float) and not isinstance(v, bool) else v)
        for h, v in raw.items()
    }


async def _write_contacts(
    session: AsyncSession,
    *,
    list_id: uuid.UUID,
    client_account_id: uuid.UUID,
    computed: ComputedImport,
) -> None:
    import json

    await session.execute(
        text("DELETE FROM contact WHERE contact_list_id = :lid"), {"lid": str(list_id)}
    )
    for r in computed.rows:
        await session.execute(
            text(
                "INSERT INTO contact (contact_list_id, client_account_id, row_index,"
                " phone_raw, phone_e164, variables, dedupe_key, diallable, exclusion_reason)"
                " VALUES (:lid, :cid, :ri, :raw, :e164, :vars, :key, :diallable, :reason)"
            ),
            {
                "lid": str(list_id),
                "cid": str(client_account_id),
                "ri": r.row_index,
                "raw": r.phone_raw,
                "e164": r.phone_e164,
                "vars": json.dumps(r.variables),
                "key": r.dedupe_key,
                "diallable": r.diallable,
                "reason": r.exclusion_reason,
            },
        )
    await session.execute(
        text("UPDATE contact_list SET diallable_count = :n WHERE id = :lid"),
        {"n": computed.diallable_count, "lid": str(list_id)},
    )


async def create_list(
    session: AsyncSession,
    *,
    client_account_id: uuid.UUID,
    agent_id: uuid.UUID,
    filename: str,
    data: bytes,
    parsed: ParsedFile,
    mapping: Mapping,
    computed: ComputedImport,
) -> uuid.UUID:
    list_id = (
        await session.execute(
            text(
                "INSERT INTO contact_list (client_account_id, agent_id, filename,"
                " row_count, column_mapping, source_file)"
                " VALUES (:cid, :aid, :fn, :rc, :map, :file) RETURNING id"
            ),
            {
                "cid": str(client_account_id),
                "aid": str(agent_id),
                "fn": filename,
                "rc": len(parsed.rows),
                "map": _mapping_to_json(mapping),
                "file": data,
            },
        )
    ).scalar_one()
    await _write_contacts(
        session,
        list_id=uuid.UUID(str(list_id)),
        client_account_id=client_account_id,
        computed=computed,
    )
    return uuid.UUID(str(list_id))


async def save_mapping(
    session: AsyncSession,
    *,
    list_id: uuid.UUID,
    client_account_id: uuid.UUID,
    mapping: Mapping,
    computed: ComputedImport,
) -> None:
    await session.execute(
        text("UPDATE contact_list SET column_mapping = :map WHERE id = :lid"),
        {"map": _mapping_to_json(mapping), "lid": str(list_id)},
    )
    await _write_contacts(
        session, list_id=list_id, client_account_id=client_account_id, computed=computed
    )


async def replace_file(
    session: AsyncSession,
    *,
    list_id: uuid.UUID,
    client_account_id: uuid.UUID,
    filename: str,
    data: bytes,
    parsed: ParsedFile,
    mapping: Mapping,
    computed: ComputedImport,
) -> None:
    await session.execute(
        text(
            "UPDATE contact_list SET filename = :fn, source_file = :file,"
            " row_count = :rc, column_mapping = :map WHERE id = :lid"
        ),
        {
            "fn": filename,
            "file": data,
            "rc": len(parsed.rows),
            "map": _mapping_to_json(mapping),
            "lid": str(list_id),
        },
    )
    await _write_contacts(
        session, list_id=list_id, client_account_id=client_account_id, computed=computed
    )


async def get_list(session: AsyncSession, list_id: uuid.UUID) -> dict[str, Any] | None:
    row = (
        await session.execute(
            text(
                "SELECT id, agent_id, filename, row_count, diallable_count,"
                " column_mapping, source_file, created_at"
                " FROM contact_list WHERE id = :lid"
            ),
            {"lid": str(list_id)},
        )
    ).first()
    if row is None:
        return None
    return {
        "id": row[0],
        "agent_id": row[1],
        "filename": row[2],
        "row_count": row[3],
        "diallable_count": row[4],
        "column_mapping": _mapping_from_json(row[5]),
        "source_file": bytes(row[6]),
        "created_at": row[7],
    }


async def list_lists(session: AsyncSession) -> list[dict[str, Any]]:
    rows = await session.execute(
        text(
            "SELECT l.id, l.filename, l.row_count, l.diallable_count, l.column_mapping,"
            " l.agent_id, a.name, l.created_at"
            " FROM contact_list l JOIN agent a ON a.id = l.agent_id"
            " ORDER BY l.created_at DESC"
        )
    )
    return [
        {
            "id": r[0],
            "filename": r[1],
            "row_count": r[2],
            "diallable_count": r[3],
            "column_mapping": _mapping_from_json(r[4]),
            "agent_id": r[5],
            "agent_name": r[6],
            "created_at": r[7],
        }
        for r in rows
    ]


async def list_health(session: AsyncSession, list_id: uuid.UUID) -> dict[str, int]:
    """Stored-row tallies for the health panel. Duplicates are not
    stored (first wins, FR-CONTACT-8): they are the gap between the
    file's row_count and the rows kept."""
    rows = await session.execute(
        text(
            "SELECT coalesce(exclusion_reason, 'ok'), count(*) FROM contact"
            " WHERE contact_list_id = :lid GROUP BY 1"
        ),
        {"lid": str(list_id)},
    )
    tally = {r[0]: int(r[1]) for r in rows}
    return {
        "ok": tally.get("ok", 0),
        "unparseable_number": tally.get("unparseable_number", 0),
        "missing_required_value": tally.get("missing_required_value", 0),
        "kept": sum(tally.values()),
    }
